#!/usr/bin/env python

"""
AUTHOR: Tiger Mou
DESCRIPTION: Reads in a CSV file of grades, config file, and graders.
             Creates a box and whisker plot, hypergeometric distribution analysis, and reformatted
             table of grades saved in an xlsx file
"""

import csv
import itertools
import json
import sys

import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter
from scipy.stats import hypergeom


def read_config(filename):
    """
    Read in config json file
    :param filename: name of the config file
    :returns: a dictionary with the config
    """
    try:
        with open(filename, 'r') as json_config:
            config = json.load(json_config)
    except EnvironmentError:
        sys.exit('EXITING... Failed to open config file.')
    return config


def read_groups(filename, columns):
    """
    Read in group map file
    (although this COULD be avoided under the assumption that each grader moves on to the
    next group after each project)
    :param filename: name of the group map file
    :param columns: array of the expected columns
    :returns: pandas table of the group map
    """
    try:
        groups_map = pd.read_csv(filename, sep=',', usecols=columns, index_col=0)
    except ValueError:
        sys.exit('EXITING... Columns specified in config not found in file %s' % filename)
    return groups_map


def read_data(filename, columns, convert):
    """
    Read in the grades data from file
    :param filename: name of the file
    :param columns: array of the expected columns
    :param convert: dictionary of converter functions
    :return: pandas table of the data
    """
    try:
        data = pd.read_csv(filename, sep=',', usecols=columns, converters=convert)
    except ValueError:
        sys.exit('EXITING... Columns specified in config not found in file %s' % filename)
    return data


def load_files():
    """
    Loads the required files from disk and returns them
    :return: config as a dictionary, groups_map as a pandas table, data as a pandas table
    """
    config_filename = sys.argv[1] if (len(sys.argv) >= 2) else 'config.json'

    # Read in config file
    config = read_config(config_filename)

    # Expected columns
    cols = config['groupHeaders']['group']
    cols = cols + [config['groupHeaders']['project']]

    # Read in groups map
    groups_map = read_groups(config['groupMapFile'], cols)

    # Expected columns
    cols = config['dataHeaders']['data']
    group = config['dataHeaders']['group']

    # Function to convert percents to floats in the 'data' columns
    def perc2float(x):
        if x == "":
            return config['convertBlanksTo']
        return float(x.strip('%')) / 100

    # turn function into dictionary from header to function
    convert = {header: perc2float for header in cols}
    cols = cols + [group]

    # Read in data file
    data = read_data(config['dataFile'], cols, convert)
    return config, groups_map, data


def create_xlsx(config, data_out, grades_keys, grades_values, bw_out):
    """
    Creates and saves the xlsx to disk, containing multiple spreadsheets with the data and analysis
    :param config: config file dictionary
    :param data_out: 2D array of data to save
    :param grades_keys: Keys of the grades
    :param grades_values: 2D array of all the grades
    :param bw_out: 2D array containing the low, mid, high to create the box plot
    :return: Does not return anything
    """
    wb = Workbook(write_only=True)

    # Probs Output
    ws1 = wb.create_sheet('Analysis Output')
    # ws1.title = 'Analysis Output'
    for row in data_out:
        ws1.append(row)
    rule_color_grad_perc = ColorScaleRule(start_type='percentile', start_value=0, start_color='3333FF',
                                     mid_type='percentile', mid_value=50, mid_color='FFFFFF',
                                     end_type='percentile', end_value=100, end_color='FF3333')

    rule_color_grad_cdfs = ColorScaleRule(start_type='num', start_value=0, start_color='3333FF',
                                          mid_type='percentile', mid_value=50, mid_color='FFFFFF',
                                          end_type='num', end_value=1, end_color='FF3333')

    # Color the Q1, Q2, Q3 and the 3 CDFs, 1 based index
    rows_percentile = ['3', '4', '5']
    rows_cdfs = ['15', '16', '17']
    last_col = get_column_letter(len(data_out[0]))

    for r in rows_percentile:
        cond_range = 'B' + r + ':' + last_col + r
        ws1.conditional_formatting.add(cond_range, rule_color_grad_perc)

    for r in rows_cdfs:
        cond_range = 'B' + r + ':' + last_col + r
        ws1.conditional_formatting.add(cond_range, rule_color_grad_cdfs)

    # Print Raw Grades by Grader data for inspection
    ws2 = wb.create_sheet(title='Raw Data By Grader')
    ws2.append(grades_keys)
    for row in [*itertools.zip_longest(*grades_values)]:
        ws2.append(row)

    rule_color_grad_grades = ColorScaleRule(start_type='min', start_color='FF8080',
                                          mid_type='percentile', mid_value=50, mid_color='FFEB84',
                                          end_type='max',  end_color='63BE7B')
    cond_range = 'A2:' + get_column_letter(len(grades_keys)) + str(max([len(i) for i in grades_values])+1)
    ws2.conditional_formatting.add(cond_range, rule_color_grad_grades)

    # Older compatible Box Whisker Plot
    ws3 = wb.create_sheet(title='Box And Whisker')
    low = bw_out[0]
    mid = np.subtract(bw_out[1], low).tolist()
    high = np.subtract(bw_out[2], bw_out[1]).tolist()
    bw_rows = [
        ["Labels"] + grades_keys,
        ["low"] + low,
        ["mid"] + mid,
        ["high"] + high
    ]
    for row in bw_rows:
        ws3.append(row)

    # Create Chart
    chart1 = BarChart(barDir='col', gapWidth='50', grouping='stacked', overlap='100')
    chart1.style = 12
    chart1.title = 'Box without Whiskers Chart'
    chart1.y_axis.title = 'Grade'
    chart1.x_axis.title = 'Grader'
    chart1.shape = 4

    chart1_data = Reference(ws3, min_col=1, min_row=2, max_row=4, max_col=len(bw_rows[0]))
    chart1_titles = Reference(ws3, min_col=2, min_row=1, max_row=1, max_col=len(bw_rows[0]))
    chart1.add_data(chart1_data, from_rows=True, titles_from_data=True)
    chart1.set_categories(chart1_titles)

    # TODO REMOVE SHADOW
    chart1.ser[0].graphicalProperties.noFill = True
    chart1.ser[0].graphicalProperties.line.noFill = True

    chart1_config = config['output']['xlsxChart']
    chart1.y_axis.scaling.min = chart1_config['y_axis']['min']
    chart1.y_axis.scaling.max = chart1_config['y_axis']['max']
    chart1.y_axis.majorUnit = chart1_config['y_axis']['unit']
    chart1.width = chart1_config['width']
    chart1.height = chart1_config['height']

    ws3.add_chart(chart1, 'A7')

    filename = config['output']['filename'] + '.xlsx'

    try:
        wb.save(filename=filename)
    except IOError:
        print("ERROR: FAILED TO SAVE XLSX! "
              "Please make sure the file is not already open", file=sys.stderr)
    else:
        print("Successfully created %s" % filename)



def main():
    config, groups_map, data = load_files()

    # Convert the 'data' df into a 2d array of grades given by each grader
    grades = {}
    combined = "ALL"
    data_cols = config['dataHeaders']['data']
    group_col = config['dataHeaders']['group']
    group_col_list = config['groupHeaders']['group']
    use_threshold = config['useAbove']
    for index1, row in data.iterrows():
        group_num = str(int(row.at[group_col]))
        for index2, item in row.filter(items=data_cols).iteritems():
            if group_num in group_col_list:
                grader = groups_map.at[index2, group_num]
                # print(group_num, "\t", index2, "\t", grader)
                # print(item)
                if item > use_threshold:
                    grades.setdefault(combined, []).append(item)
                    grades.setdefault(grader, []).append(item)

    # Create box plot
    grades_values = list(grades.values())
    grades_keys = list(grades.keys())
    try:
        box_values = plt.boxplot(grades_values, labels=grades_keys)
    except ValueError as e:
        print('Keys: %a' % grades_keys, file=sys.stderr)
        print('Values: %a' % grades_values, file=sys.stderr)
        sys.exit('EXCEPTION: ValueError! ' + str(e))
    # Retrieve data from the box plot
    res = {key: [v.get_ydata() for v in value] for key, value in box_values.items()}

    # EXTRACT TABLE DATA
    whiskers_min = [min(item) for item in res['whiskers'][::2]]  # Lower Whisker
    whiskers_Q1 = [max(item) for item in res['whiskers'][::2]]  # Q1
    whiskers_Q3 = [min(item) for item in res['whiskers'][1::2]]  # Q3
    whiskers_max = [max(item) for item in res['whiskers'][1::2]]  # Upper Whisker
    medians = [item[0] for item in res['medians']]  # Q2

    # ALTERNATIVE METHODS, EQUIVALENT DATA
    # caps_lower = [item[0] for item in res['caps'][::2]] # Lower Whisker
    # caps_upper = [item[0] for item in res['caps'][1::2]] # Upper Whisker
    # boxes_lower = [min(item) for item in res['boxes']] # Q1
    # boxes_upper = [max(item) for item in res['boxes']] # Q3

    # print(res['fliers']) # outliers
    # print(res['means']) # EMPTY ARRAY!

    # Format and display box plot
    if config['output']['pyplot']:
        # TODO CONFIG FLAGS
        plt.grid(axis='y')
        plt.figure(figsize=(12, 4))
        plt.yticks(np.arange(use_threshold, 1.1, step=0.05))
        try:
            plt.savefig('output_box_plot.png', dpi=200)
        except IOError:
            print("ERROR: FAILED TO SAVE PYPLOT FIGURE! "
                  "Please make sure the file is not already open", file=sys.stderr)
        # plt.show()

    # Calculate hypergeometric CDFs
    h_cuts = [whiskers_Q1, medians, whiskers_Q3]
    M = len(grades[combined])  # total overall
    N = [len(graded) for graded in grades_values]  # number the grader completed

    # CALCULATE n,x using >= cut value to find chance of getting as extreme or more extreme

    # number of total satisfying. (Matrix len(grades) by len(h_cuts)
    n = [[sum([1 for item in grades[combined] if item >= val]) for val in cut]
         for cut in h_cuts]
    # number of selected satisfying. (Matrix len(grades) by len(h_cuts)
    x = [[sum([1 for item in gr if item >= val])
          for val, gr in zip(cut, grades_values)]
         for cut in h_cuts]
    # Prob to get sample LESS 'extreme' (with more selected<cut). (Matrix len(grades) by len(h_cuts)
    h_cuts_probs = [[hypergeom.cdf(c_x, M, c_n, c_N)
                     for c_x, c_n, c_N in zip(cut_x, cut_n, N)]
                    for cut, cut_n, cut_x in zip(h_cuts, n, x)]  #

    data_out = [
        ["Labels"] + grades_keys,
        ["whisker_min"] + whiskers_min,
        ["Q1"] + whiskers_Q1,
        ["Q2"] + medians,
        ["Q3"] + whiskers_Q3,
        ["whisker_max"] + whiskers_max,
        ["M"] + [M for _ in range(len(N))],
        ["N"] + N,
        ["n_Q1"] + n[0],
        ["n_Q2"] + n[1],
        ["n_Q3"] + n[2],
        ["x_Q1"] + x[0],
        ["x_Q2"] + x[1],
        ["x_Q3"] + x[2],
        ["Fx(x)_Q1"] + h_cuts_probs[0],
        ["Fx(x)_Q2"] + h_cuts_probs[1],
        ["Fx(x)_Q3"] + h_cuts_probs[2]
    ]

    if config['output']['format'] == 'xlsx':
        create_xlsx(config, data_out,
                    grades_keys, grades_values, [whiskers_Q1, medians, whiskers_Q3])
    else:
        with open(config['output']['filename'] + '.csv', 'w+') as csv_outfile:
            csv_w = csv.writer(csv_outfile, delimiter=',', lineterminator='\n')
            csv_w.writerows(data_out)


if __name__ == '__main__':
    main()
    print('Done.')
